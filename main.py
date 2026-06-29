"""
بوت تليغرام لإدارة محل الكوزمتك - نسخة Render (24/7 مجاني)
"""

import os, json, logging, tempfile, io, base64
from datetime import datetime

from google import genai
from google.genai import types
from telegram import Update
from telegram.ext import Application, MessageHandler, CommandHandler, filters, ContextTypes
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload
import openpyxl

# ═══════════════════════════════════════════════════
TELEGRAM_TOKEN       = os.environ["TELEGRAM_TOKEN"]
GOOGLE_FILE_ID       = os.environ["GOOGLE_FILE_ID"]
GEMINI_API_KEY       = os.environ["GEMINI_API_KEY"]
SERVICE_ACCOUNT_JSON = os.environ["SERVICE_ACCOUNT_JSON"]
# ═══════════════════════════════════════════════════

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(message)s')
logger = logging.getLogger(__name__)

client = genai.Client(api_key=GEMINI_API_KEY)

SYSTEM_PROMPT = """You manage a cosmetics store. Analyze Arabic text or image and return ONLY a JSON array.

Actions:
1. Purchase/اشتريت/شراء: {"action":"add_purchase","data":{"product":"name","category":"type","supplier":"","qty":NUMBER,"price":NUMBER},"message":"✅ تم تسجيل الشراء"}
2. Sale/بعت/بيع: {"action":"add_sale","data":{"product":"name","qty":NUMBER,"sell_price":NUMBER},"message":"✅ تم تسجيل البيع"}
3. New product/أضف منتج: {"action":"add_product","data":{"name":"","category":"عام","qty":0},"message":"✅ تم إضافة المنتج"}
4. Query/كم/استعلام: {"action":"query","data":{"product":"name"},"message":""}

For images with tables: extract ALL rows as separate items.
Return JSON array only. No markdown. No explanation."""

# ─── Google Drive ───────────────────────────────
def get_drive_service():
    info = json.loads(SERVICE_ACCOUNT_JSON)
    creds = Credentials.from_service_account_info(
        info, scopes=["https://www.googleapis.com/auth/drive"]
    )
    return build("drive", "v3", credentials=creds)

def download_excel():
    service = get_drive_service()
    req = service.files().get_media(fileId=GOOGLE_FILE_ID)
    buf = io.BytesIO()
    dl = MediaIoBaseDownload(buf, req)
    done = False
    while not done:
        _, done = dl.next_chunk()
    buf.seek(0)
    return buf

def upload_excel(file_bytes):
    service = get_drive_service()
    tmp_path = os.path.join(tempfile.gettempdir(), "cosmetic_upload.xlsx")
    with open(tmp_path, "wb") as f:
        f.write(file_bytes)
    media = MediaFileUpload(tmp_path, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    service.files().update(fileId=GOOGLE_FILE_ID, media_body=media).execute()
    try:
        os.unlink(tmp_path)
    except:
        pass

# ─── Gemini ─────────────────────────────────────
def call_gemini(text=None, image_bytes=None):
    try:
        if image_bytes:
            import PIL.Image
            img = PIL.Image.open(io.BytesIO(image_bytes))
            img_buf = io.BytesIO()
            img.save(img_buf, format='JPEG')
            prompt = SYSTEM_PROMPT + "\nExtract ALL rows from this image as JSON array."
            resp = client.models.generate_content(
                model="gemini-2.5-flash",
                contents=[
                    types.Part.from_bytes(data=img_buf.getvalue(), mime_type="image/jpeg"),
                    types.Part.from_text(text=prompt)
                ]
            )
        else:
            prompt = SYSTEM_PROMPT + f"\nUser: {text}"
            resp = client.models.generate_content(
                model="gemini-2.5-flash",
                contents=prompt
            )

        raw = resp.text.strip().replace("```json","").replace("```","").strip()
        arr_s = raw.find("["); arr_e = raw.rfind("]") + 1
        obj_s = raw.find("{"); obj_e = raw.rfind("}") + 1

        if arr_s >= 0 and arr_e > arr_s:
            result = json.loads(raw[arr_s:arr_e])
            return result if isinstance(result, list) else [result]
        if obj_s >= 0 and obj_e > obj_s:
            return [json.loads(raw[obj_s:obj_e])]

        return [{"action":"error","data":{},"message":"⚠️ لم أفهم الطلب"}]
    except Exception as e:
        logger.error(f"Gemini error: {e}")
        return [{"action":"error","data":{},"message":f"❌ خطأ: {str(e)[:150]}"}]

# ─── Excel helpers ───────────────────────────────
def find_product_row(ws, name, max_row=102):
    name = name.strip()
    for r in range(3, max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val and name in str(val).strip():
            return r
    return -1

def find_empty_row(ws, name_col=2, start=3, max_r=202):
    for r in range(start, max_r + 1):
        if not ws.cell(row=r, column=1).value and not ws.cell(row=r, column=name_col).value:
            return r
    return max_r

def get_or_create_product(ws_inv, name, category="عام"):
    row = find_product_row(ws_inv, name)
    if row > 0:
        return row
    row = find_empty_row(ws_inv, name_col=1, start=3, max_r=102)
    ws_inv.cell(row=row, column=1, value=name)
    ws_inv.cell(row=row, column=2, value=category)
    ws_inv.cell(row=row, column=3, value=0)
    return row

# ─── تطبيق الأوامر ───────────────────────────────
def apply_single(wb, cmd):
    action = cmd.get("action","")
    data   = cmd.get("data", {})
    today  = datetime.now().strftime("%Y/%m/%d")
    ws_inv = wb["المخزون"]

    if action == "add_purchase":
        product  = data.get("product","")
        category = data.get("category","عام")
        qty      = int(data.get("qty", 0))
        price    = float(data.get("price", 0))
        supplier = data.get("supplier","")

        ws_pur = wb["المشتريات"]
        row = find_empty_row(ws_pur, name_col=2)
        ws_pur.cell(row=row, column=1, value=today)
        ws_pur.cell(row=row, column=2, value=product)
        ws_pur.cell(row=row, column=3, value=supplier)
        ws_pur.cell(row=row, column=4, value=qty)
        ws_pur.cell(row=row, column=5, value=price)
        ws_pur.cell(row=row, column=6, value=f"=D{row}*E{row}")

        inv_row = get_or_create_product(ws_inv, product, category)
        old_qty = int(ws_inv.cell(row=inv_row, column=3).value or 0)
        ws_inv.cell(row=inv_row, column=3, value=old_qty + qty)
        return f"✅ تم تسجيل شراء {qty} {product}\n📦 المخزون: {old_qty} ← {old_qty + qty}"

    elif action == "add_sale":
        product    = data.get("product","")
        qty        = int(data.get("qty", 0))
        sell_price = float(data.get("sell_price", 0))

        ws_sal = wb["المبيعات"]
        row = find_empty_row(ws_sal, name_col=2)
        ws_sal.cell(row=row, column=1, value=today)
        ws_sal.cell(row=row, column=2, value=product)
        ws_sal.cell(row=row, column=3, value=qty)
        ws_sal.cell(row=row, column=4, value=sell_price)
        ws_sal.cell(row=row, column=5, value=f"=C{row}*D{row}")

        inv_row = find_product_row(ws_inv, product)
        if inv_row > 0:
            old_qty = int(ws_inv.cell(row=inv_row, column=3).value or 0)
            new_qty = max(0, old_qty - qty)
            ws_inv.cell(row=inv_row, column=3, value=new_qty)
            return f"✅ تم تسجيل بيع {qty} {product}\n📦 المخزون: {old_qty} ← {new_qty}"
        else:
            new_row = find_empty_row(ws_inv, name_col=1, start=3, max_r=102)
            ws_inv.cell(row=new_row, column=1, value=product)
            ws_inv.cell(row=new_row, column=2, value="عام")
            ws_inv.cell(row=new_row, column=3, value=0)
            return f"⚠️ تم تسجيل البيع لكن '{product}' لم يكن في المخزون"

    elif action == "add_product":
        name     = data.get("name","")
        category = data.get("category","عام")
        qty      = int(data.get("qty", 0))
        inv_row  = find_product_row(ws_inv, name)
        if inv_row < 0:
            inv_row = find_empty_row(ws_inv, name_col=1, start=3, max_r=102)
            ws_inv.cell(row=inv_row, column=1, value=name)
            ws_inv.cell(row=inv_row, column=2, value=category)
            ws_inv.cell(row=inv_row, column=3, value=qty)
        return f"✅ تم إضافة {name} للمخزون"

    elif action == "query":
        product = data.get("product","")
        inv_row = find_product_row(ws_inv, product)
        if inv_row > 0:
            qty      = ws_inv.cell(row=inv_row, column=3).value or 0
            category = ws_inv.cell(row=inv_row, column=2).value or ""
            status   = "🔴 نفد" if qty == 0 else ("🟡 منخفض" if qty <= 5 else "🟢 متوفر")
            return f"📦 {product}\nالنوع: {category}\nالكمية: {qty} {status}"
        return f"⚠️ '{product}' غير موجود في المخزون"

    elif action == "error":
        return cmd.get("message","⚠️ خطأ")

    return cmd.get("message","✅ تم")

def apply_all(wb, cmds):
    if isinstance(cmds, dict): cmds = [cmds]
    results = [apply_single(wb, c) for c in cmds if c]
    return "\n\n".join(r for r in results if r)

# ─── Telegram Handlers ───────────────────────────
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "🧴 *بوت محل الكوزمتك جاهز!*\n\n"
        "📝 اكتب مثلاً:\n"
        "• اشتريت 20 شامبو بسعر 3000\n"
        "• بعت 5 كريم بسعر 5000\n"
        "• كم كمية الشامبو؟\n\n"
        "📸 أو صوّر الفاتورة",
        parse_mode="Markdown"
    )

async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    msg = await update.message.reply_text("⏳ جاري المعالجة...")
    try:
        cmds = call_gemini(text=update.message.text)
        if not isinstance(cmds, list): cmds = [cmds]
        if all(c.get("action") in ["query","error"] for c in cmds):
            buf = download_excel()
            wb  = openpyxl.load_workbook(buf)
            await msg.edit_text(apply_all(wb, cmds))
            return
        buf = download_excel()
        wb  = openpyxl.load_workbook(buf)
        result = apply_all(wb, cmds)
        out = io.BytesIO(); wb.save(out)
        upload_excel(out.getvalue())
        await msg.edit_text(result)
    except Exception as e:
        logger.error(e)
        await msg.edit_text(f"❌ خطأ: {e}")

async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    msg = await update.message.reply_text("📸 جاري تحليل الصورة...")
    try:
        pf = await context.bot.get_file(update.message.photo[-1].file_id)
        pb = bytes(await pf.download_as_bytearray())
        cmds = call_gemini(image_bytes=pb)
        if not isinstance(cmds, list): cmds = [cmds]
        buf = download_excel(); wb = openpyxl.load_workbook(buf)
        result = apply_all(wb, cmds)
        out = io.BytesIO(); wb.save(out); upload_excel(out.getvalue())
        await msg.edit_text(result)
    except Exception as e:
        await msg.edit_text(f"❌ خطأ: {e}")

def main():
    app = Application.builder().token(TELEGRAM_TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))
    app.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    logger.info("🤖 البوت يعمل على Render!")
    app.run_polling()

if __name__ == "__main__":
    main()
